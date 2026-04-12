"""
The FullPicture — Flask backend
Stores user logins, quiz results, comments, article views, and submissions
in an Excel workbook on the local machine.
"""

import datetime
import hashlib
import json
import os
import re
import shutil
from pathlib import Path

from flask import Flask, jsonify, request, send_from_directory, session
from flask_cors import CORS
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from werkzeug.security import check_password_hash, generate_password_hash

app = Flask(__name__, static_folder=".")
app.secret_key = os.environ.get("SECRET_KEY", "fullpicture-secret-2026")
CORS(app, supports_credentials=True)

BASE_DIR = Path(__file__).resolve().parent
DESKTOP_EXCEL_DIR = Path.home() / "Desktop" / "Excel"
EXCEL_FILE = DESKTOP_EXCEL_DIR / "fullpicture_data.xlsx"
LEGACY_EXCEL_FILE = BASE_DIR / "fullpicture_data.xlsx"

EMAIL_RE = re.compile(r"^[A-Z0-9._%+-]+@[A-Z0-9.-]+\.[A-Z]{2,}$", re.I)
USERNAME_RE = re.compile(r"^[A-Za-z][A-Za-z0-9_-]{2,23}$")
SPECIAL_RE = re.compile(r"[^A-Za-z0-9]")
HEX64_RE = re.compile(r"^[a-f0-9]{64}$")

ALLOWED_IDEOLOGIES = {"left", "center", "right"}
ALLOWED_LEANS = {"left", "center", "right"}
MAX_COMMENT_LENGTH = 2000
MAX_SUBMISSION_DESCRIPTION = 4000

SHEETS = {
    "Users": [
        "id",
        "username",
        "email",
        "password_hash",
        "created_at",
        "last_login",
        "ideology",
        "quiz_scores",
    ],
    "Sessions": [
        "session_id",
        "user_id",
        "username",
        "login_time",
        "logout_time",
        "ip_address",
    ],
    "Comments": [
        "id",
        "article_id",
        "article_title",
        "user_id",
        "username",
        "comment_text",
        "posted_at",
    ],
    "Views": ["id", "article_id", "article_title", "user_id", "username", "viewed_at"],
    "Submissions": [
        "id",
        "title",
        "url",
        "author",
        "lean",
        "topic",
        "description",
        "submitted_by",
        "submitted_at",
        "status",
    ],
}

HEADER_FILL = PatternFill("solid", fgColor="15304F")
HEADER_FONT = Font(color="FFFFFF", bold=True)


def now():
    return datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")


def format_short_date(ts):
    try:
        return datetime.datetime.strptime(ts, "%Y-%m-%d %H:%M:%S").strftime("%d %b")
    except Exception:
        return ts


def ensure_storage_location():
    DESKTOP_EXCEL_DIR.mkdir(parents=True, exist_ok=True)
    if not EXCEL_FILE.exists() and LEGACY_EXCEL_FILE.exists():
        shutil.copy2(LEGACY_EXCEL_FILE, EXCEL_FILE)


def init_excel():
    ensure_storage_location()
    if EXCEL_FILE.exists():
        wb = load_workbook(EXCEL_FILE)
    else:
        wb = Workbook()
        wb.remove(wb.active)

    for sheet_name, cols in SHEETS.items():
        if sheet_name not in wb.sheetnames:
            ws = wb.create_sheet(sheet_name)
            for i, col in enumerate(cols, 1):
                cell = ws.cell(row=1, column=i, value=col.replace("_", " ").title())
                cell.font = HEADER_FONT
                cell.fill = HEADER_FILL
                cell.alignment = Alignment(horizontal="center")
            ws.freeze_panes = "A2"

    wb.save(EXCEL_FILE)


def get_wb():
    ensure_storage_location()
    return load_workbook(EXCEL_FILE)


def save_wb(wb):
    ensure_storage_location()
    wb.save(EXCEL_FILE)


def next_id(ws):
    return ws.max_row


def hash_pw(password):
    return generate_password_hash(password)


def verify_pw(stored_hash, password):
    if not stored_hash:
        return False
    if HEX64_RE.fullmatch(stored_hash):
        return stored_hash == hashlib.sha256(password.encode()).hexdigest()
    return check_password_hash(stored_hash, password)


def upgrade_legacy_hash_if_needed(ws, row_num, stored_hash, password):
    if HEX64_RE.fullmatch(stored_hash):
        ws.cell(row_num, 4).value = hash_pw(password)


def clean_text(value, limit=None):
    text = str(value or "").strip()
    if limit is not None:
        text = text[:limit]
    return text


def has_simple_pattern(password):
    lowered = password.lower()
    weak_fragments = ["password", "qwerty", "admin", "welcome", "1234", "abcd"]
    if any(fragment in lowered for fragment in weak_fragments):
        return True

    sequences = [
        "abcdefghijklmnopqrstuvwxyz",
        "0123456789",
    ]
    for sequence in sequences:
        for idx in range(len(sequence) - 3):
            snippet = sequence[idx : idx + 4]
            if snippet in lowered or snippet[::-1] in lowered:
                return True

    max_repeat = 1
    current_repeat = 1
    for idx in range(1, len(password)):
        if password[idx] == password[idx - 1]:
            current_repeat += 1
            max_repeat = max(max_repeat, current_repeat)
        else:
            current_repeat = 1
    return max_repeat >= 4


def password_requirements(password, username="", email=""):
    email_local = email.split("@")[0].lower() if "@" in email else ""
    lowered = password.lower()

    return [
        (len(password) >= 14, "Password must be at least 14 characters long."),
        (any(ch.islower() for ch in password), "Password must include a lowercase letter."),
        (any(ch.isupper() for ch in password), "Password must include an uppercase letter."),
        (any(ch.isdigit() for ch in password), "Password must include a number."),
        (bool(SPECIAL_RE.search(password)), "Password must include a special character."),
        (" " not in password, "Password cannot contain spaces."),
        (username.lower() not in lowered if username else True, "Password cannot contain your username."),
        (email_local not in lowered if email_local else True, "Password cannot contain your email name."),
        (not has_simple_pattern(password), "Password cannot use obvious sequences or repeated characters."),
    ]


def validate_registration(username, email, password):
    errors = []

    if not USERNAME_RE.fullmatch(username):
        errors.append(
            "Username must start with a letter, be 3-24 characters, and use only letters, numbers, underscores, or hyphens."
        )

    if not EMAIL_RE.fullmatch(email):
        errors.append("Enter a valid email address.")

    for passed, message in password_requirements(password, username=username, email=email):
        if not passed:
            errors.append(message)

    return errors


def _log_session(uid, username):
    wb = get_wb()
    ws = wb["Sessions"]
    sid = next_id(ws)
    ip_address = request.remote_addr or "unknown"
    ws.append([sid, uid, username, now(), "", ip_address])
    save_wb(wb)


@app.route("/api/register", methods=["POST"])
def register():
    data = request.get_json(silent=True) or {}
    username = clean_text(data.get("username"), 24)
    email = clean_text(data.get("email"), 255).lower()
    password = data.get("password", "")
    ideology = clean_text(data.get("ideology"), 20).lower()
    if ideology not in ALLOWED_IDEOLOGIES:
        ideology = ""

    quiz_scores = data.get("quiz_scores", {})
    errors = validate_registration(username, email, password)
    if errors:
        return jsonify({"ok": False, "msg": errors[0], "errors": errors}), 400

    wb = get_wb()
    ws = wb["Users"]
    for row in range(2, ws.max_row + 1):
        row_username = (ws.cell(row, 2).value or "").strip().lower()
        row_email = (ws.cell(row, 3).value or "").strip().lower()
        if row_email == email:
            return jsonify({"ok": False, "msg": "Email already registered."}), 409
        if row_username == username.lower():
            return jsonify({"ok": False, "msg": "Username already taken."}), 409

    uid = next_id(ws)
    ws.append(
        [
            uid,
            username,
            email,
            hash_pw(password),
            now(),
            now(),
            ideology,
            json.dumps(quiz_scores, ensure_ascii=True),
        ]
    )
    save_wb(wb)

    session["user_id"] = uid
    session["username"] = username
    session["ideology"] = ideology
    _log_session(uid, username)

    return jsonify(
        {"ok": True, "user": {"id": uid, "username": username, "ideology": ideology}}
    )


@app.route("/api/login", methods=["POST"])
def login():
    data = request.get_json(silent=True) or {}
    email = clean_text(data.get("email"), 255).lower()
    password = data.get("password", "")

    if not email or not password:
        return jsonify({"ok": False, "msg": "Email and password are required."}), 400

    wb = get_wb()
    ws = wb["Users"]
    for row in range(2, ws.max_row + 1):
        row_email = (ws.cell(row, 3).value or "").strip().lower()
        stored_hash = ws.cell(row, 4).value or ""
        if row_email == email and verify_pw(stored_hash, password):
            uid = ws.cell(row, 1).value
            username = ws.cell(row, 2).value
            ideology = (ws.cell(row, 7).value or "").strip().lower()

            upgrade_legacy_hash_if_needed(ws, row, stored_hash, password)
            ws.cell(row, 6).value = now()
            save_wb(wb)

            session["user_id"] = uid
            session["username"] = username
            session["ideology"] = ideology
            _log_session(uid, username)

            return jsonify(
                {"ok": True, "user": {"id": uid, "username": username, "ideology": ideology}}
            )

    return jsonify({"ok": False, "msg": "Invalid email or password."}), 401


@app.route("/api/logout", methods=["POST"])
def logout():
    session.clear()
    return jsonify({"ok": True})


@app.route("/api/me", methods=["GET"])
def me():
    if "user_id" not in session:
        return jsonify({"ok": False, "user": None})
    return jsonify(
        {
            "ok": True,
            "user": {
                "id": session["user_id"],
                "username": session["username"],
                "ideology": session.get("ideology", ""),
            },
        }
    )


@app.route("/api/ideology", methods=["POST"])
def save_ideology():
    data = request.get_json(silent=True) or {}
    ideology = clean_text(data.get("ideology"), 20).lower()
    if ideology not in ALLOWED_IDEOLOGIES:
        return jsonify({"ok": False, "msg": "Invalid ideology."}), 400

    quiz_scores = json.dumps(data.get("quiz_scores", {}), ensure_ascii=True)
    session["ideology"] = ideology

    if "user_id" in session:
        wb = get_wb()
        ws = wb["Users"]
        user_id = session["user_id"]
        for row in range(2, ws.max_row + 1):
            if ws.cell(row, 1).value == user_id:
                ws.cell(row, 7).value = ideology
                ws.cell(row, 8).value = quiz_scores
                break
        save_wb(wb)

    return jsonify({"ok": True})


@app.route("/api/comments/<int:article_id>", methods=["GET"])
def get_comments(article_id):
    wb = get_wb()
    ws = wb["Comments"]
    comments = []
    for row in range(2, ws.max_row + 1):
        if ws.cell(row, 2).value == article_id:
            posted_at = ws.cell(row, 7).value or ""
            comments.append(
                {
                    "id": ws.cell(row, 1).value,
                    "author": ws.cell(row, 5).value,
                    "text": ws.cell(row, 6).value,
                    "date": format_short_date(posted_at),
                }
            )
    return jsonify(comments)


@app.route("/api/comments", methods=["POST"])
def post_comment():
    data = request.get_json(silent=True) or {}
    article_id = data.get("article_id")
    article_title = clean_text(data.get("article_title"), 255)
    text = clean_text(data.get("text"), MAX_COMMENT_LENGTH)
    username = clean_text(data.get("username"), 40) or session.get("username", "Anonymous")
    user_id = session.get("user_id", 0)

    if not isinstance(article_id, int):
        return jsonify({"ok": False, "msg": "Article id is required."}), 400
    if not text:
        return jsonify({"ok": False, "msg": "Comment cannot be empty."}), 400

    wb = get_wb()
    ws = wb["Comments"]
    comment_id = next_id(ws)
    timestamp = now()
    ws.append([comment_id, article_id, article_title, user_id, username, text, timestamp])
    save_wb(wb)

    return jsonify(
        {
            "ok": True,
            "comment": {
                "id": comment_id,
                "author": username,
                "text": text,
                "date": format_short_date(timestamp),
            },
        }
    )


@app.route("/api/view", methods=["POST"])
def log_view():
    data = request.get_json(silent=True) or {}
    article_id = data.get("article_id")
    if not isinstance(article_id, int):
        return jsonify({"ok": False, "msg": "Article id is required."}), 400

    wb = get_wb()
    ws = wb["Views"]
    view_id = next_id(ws)
    ws.append(
        [
            view_id,
            article_id,
            clean_text(data.get("article_title"), 255),
            session.get("user_id", 0),
            session.get("username", "guest"),
            now(),
        ]
    )
    save_wb(wb)
    return jsonify({"ok": True})


@app.route("/api/submit", methods=["POST"])
def submit_article():
    data = request.get_json(silent=True) or {}
    title = clean_text(data.get("title"), 255)
    url = clean_text(data.get("url"), 500)
    author = clean_text(data.get("author"), 120)
    lean = clean_text(data.get("lean"), 20).lower()
    topic = clean_text(data.get("topic"), 40).lower()
    description = clean_text(data.get("description"), MAX_SUBMISSION_DESCRIPTION)

    if not title or not url or not topic or lean not in ALLOWED_LEANS:
        return jsonify({"ok": False, "msg": "Complete the required submission fields."}), 400

    if not (url.startswith("http://") or url.startswith("https://")):
        return jsonify({"ok": False, "msg": "Article URL must start with http:// or https://."}), 400

    wb = get_wb()
    ws = wb["Submissions"]
    submission_id = next_id(ws)
    ws.append(
        [
            submission_id,
            title,
            url,
            author,
            lean,
            topic,
            description,
            session.get("username", "anonymous"),
            now(),
            "Pending",
        ]
    )
    save_wb(wb)
    return jsonify({"ok": True})


@app.route("/api/admin/stats", methods=["GET"])
def admin_stats():
    wb = get_wb()
    return jsonify(
        {
            "users": max(0, wb["Users"].max_row - 1),
            "comments": max(0, wb["Comments"].max_row - 1),
            "views": max(0, wb["Views"].max_row - 1),
            "submissions": max(0, wb["Submissions"].max_row - 1),
            "storage_path": str(EXCEL_FILE),
        }
    )


@app.route("/")
def index():
    return send_from_directory(".", "index.html")


if __name__ == "__main__":
    init_excel()
    print("\nThe FullPicture backend is running.")
    print(f"Data stored in: {EXCEL_FILE}")
    print("Open http://localhost:5000 in your browser.\n")
    app.run(debug=True, port=5000)
